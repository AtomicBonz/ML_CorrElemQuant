# -*- coding: utf-8 -*-
"""
Created on Fri, 5 May 2022, 13:0 0h

Feature dimensional reduction:
   1) Bifurcation: train set / test set
   2 )Encapsulation of MLtrain & MLtest routines
   3) Fit unpruned feature set to test set --> report MAPE
   4) Spearman prune, then fit to reduced feature set --> report MAPE
   5) Randseed perturubation prune, then fit to 2' reduced feature set --> report MAPE
   6) Test on 2' pruned feature set. 
   7) Plots: 
        Scatterplot of train (DOES NOT TEST)
        Histogram of errors
        Barplot of feature ordinal rankings < Randomness Feature 
        Barplot comparison of MAPEs between: unpruned feature set, Spearman-pruned only, Randseed perturbation pruned 

@author: bgoh, JHS
"""
#%% (1) Import all necessary Packages
# import sys
#sys.modules[__name__].__dict__.clear() #start with a clean workspace
from __future__ import unicode_literals #you need this line to use greek letters
import time
startTime = time.time()

# import re
import numpy as np
import pandas as pd
# import pymatgen as mg
from sklearn.preprocessing import StandardScaler
# from sklearn import metrics
# from sklearn.linear_model import LinearRegression
# from sklearn import preprocessing
from sklearn import model_selection
from sklearn.ensemble import RandomForestRegressor

import matplotlib.pyplot as plt
import sklearn.model_selection as xval
# from sklearn import datasets
# from sklearn import decomposition
# from sklearn.covariance import EmpiricalCovariance
from sklearn.metrics import mean_squared_error
from math import sqrt

import pathlib


from scipy import stats
from openpyxl import Workbook
import shap

# from itertools import combinations

#%% (2) Input files & specification params
normScale = StandardScaler()
dataFile = 'ML_InpAsOf_2022_12_20_ICPMS_all.xlsx'
trainData_whichSheet = 'Sheet_manSpearPrune(6)' #you need 

#testData_whichSheet = 'Pl23_EBSDtextur'

#testSheet = 'ML_test'
randSeedVal = 1
modelType = "theUshe_"
targetVariableType = 'saltConc_ppm_ICPMS_'
unit = "[ppm]"
normScale = StandardScaler()#it's a normscaler
threshold_spear = 0.97 #Spearman cutoff threshold
randSeedPerturbReps = 100
SS316CorrCriteria = 1000 #[ppm]
featureGroupRecord = modelType + "featureGroups.xlsx"
resultsExportDir = 'LatestOutput/shaps/'

elems = ["total"]#["Cr", "Fe","Mn","Ni","total"] #

df_train = pd.read_excel(dataFile, sheet_name = trainData_whichSheet)#),sheet_name="Sheet3")
df_train = df_train.dropna(axis=0)

#the scaled set
df_colnames = df_train.columns
df_train_scaled = normScale.fit_transform(df_train)
df_train_scaled = pd.DataFrame(df_train_scaled,columns = df_colnames)

# df_test = pd.read_excel(dataFile, sheet_name = testData_whichSheet)#),sheet_name="Sheet3")
# df_test = df_test.dropna(axis=0)

#%% (3) Homebrew functions
"""
Define the function that will interrogate the individual trees in the RF
jrhs wrote this function
"""


def pred_ints(model, X, percentile=95):
    err_down = []
    err_up = []
    for x in range(len(X)):
        preds = []
        for pred in model.estimators_:
            preds.append(pred.predict(X)[x])
        err_down.append(np.percentile(preds, (100 - percentile) / 2. ))
        err_up.append(np.percentile(preds, 100 - (100 - percentile) / 2.))
    return err_down, err_up
    
def MLtrainsetMetrics(model,X_inps,y_inps):
    X_inps_np = X_inps.to_numpy()
    y_inps_np = y_inps.to_numpy()
    model.fit(X_inps_np,y_inps_np)
    Rsq = model.score(X_inps_np,y_inps_np,sample_weight=None)
    y_fit = model.predict(X_inps_np)
    RMSE_fit = sqrt(mean_squared_error(y_inps_np, y_fit))
    
    uncert_down_y_predict, uncert_up_y_predict = pred_ints(model,X_inps_np, percentile=90)
    uncert_down_y_predict_np = np.array(uncert_down_y_predict)
    uncert_up_y_predict_np = np.array(uncert_up_y_predict)
    uncertMargin_y_pred = (uncert_up_y_predict_np - uncert_down_y_predict_np)/2
	
    yfit_xvalPredict = model_selection.cross_val_predict(model,X_inps_np,y_inps_np,cv=5)
    RMSE_xval = sqrt(mean_squared_error(y_inps_np, yfit_xvalPredict))
    y_inps_arr = np.array(y_inps)
    yfit_xvalPredict_arr = np.array(yfit_xvalPredict)
    yfit_absPredErrs = abs(yfit_xvalPredict_arr - y_inps_arr)
    yfit_percPredErrs = np.divide(yfit_absPredErrs,abs(y_inps_arr))*100 #y prediction errors, units [%]
    yfit_xvalMAPE = yfit_percPredErrs.mean()
    yfit_xvalMAPE_stdev = yfit_percPredErrs.std()
    featImpts = model.feature_importances_
    return Rsq, y_fit, RMSE_fit, uncert_down_y_predict_np, uncert_up_y_predict_np, uncertMargin_y_pred, yfit_xvalPredict, RMSE_xval, yfit_percPredErrs, yfit_xvalMAPE, yfit_xvalMAPE_stdev, featImpts

def MLtestsetMetrics(model,X_inps_train,y_inps_train,X_inps_test,y_inps_test):
    X_inps_train_np = X_inps_train.to_numpy()
    y_inps_train_np = y_inps_train.to_numpy()
    model.fit(X_inps_train_np,y_inps_train_np)
    X_inps_test_np = X_inps_test.to_numpy()
    y_test_pred = model.predict(X_inps_test_np)
    Rsq = model.score(X_inps_train,y_inps_train,sample_weight=None)
    #y_fit = model.predict(X_inps)
    #RMSE_fit = sqrt(mean_squared_error(y_inps, y_fit))
    
    #uncert_down_y_predict, uncert_up_y_predict = pred_ints(model,X_inps, percentile=90)
    #uncert_down_y_predict_np = np.array(uncert_down_y_predict)
    #uncert_up_y_predict_np = np.array(uncert_up_y_predict)
    #uncertMargin_y_pred = (uncert_up_y_predict_np - uncert_down_y_predict_np)/2
	
    #yfit_xvalPredict = model_selection.cross_val_predict(model,X_inps,y_inps,cv=5)
    #RMSE_xval = sqrt(mean_squared_error(y_inps, yfit_xvalPredict))
    #y_inps_arr = np.array(y_inps)
    y_fit_arr = np.array(y_test_pred)
    yfit_absPredErrs = abs(y_fit_arr - y_inps_test)
    yfit_percPredErrs = np.divide(yfit_absPredErrs,abs(y_inps_test))*100 #y prediction errors, units [%]
    yfit_MAPE = yfit_percPredErrs.mean()
    yfit_MAPE_stdev = yfit_percPredErrs.std()
    return  y_test_pred, yfit_percPredErrs, yfit_MAPE, yfit_MAPE_stdev,Rsq

#%%
def scatter_hist(x, y,x_featureName):
    # definitions for the axes
    left, width = 0.1, 0.65
    bottom, height = 0.1, 0.65
    spacing = 0.03
    
    
    rect_scatter = [left, bottom + 0.2+spacing, width, height]
    rect_histx = [left, bottom, width, 0.2]
    #rect_histx = [left, bottom + height + spacing, width, 0.2]
    #rect_histy = [left + width + spacing, bottom, 0.2, height]
    
    # start with a square Figure
    fig = plt.figure(figsize=(8, 8))
    
    ax = fig.add_axes(rect_scatter)
    ax_histx = fig.add_axes(rect_histx, sharex=ax)
    ax_histx.patch.set_alpha(0)
    #ax_histy = fig.add_axes(rect_histy, sharey=ax)
    # no labels
    ax_histx.tick_params(axis="x",labelleft=True)
    #ax_histy.tick_params(axis="y", labelleft=False)

    # the scatter plot:
    ax.scatter(x, y)
    plt.xlabel("population relative " + x_featureName)
    plt.text(-2.6,80,"Shapley Value",rotation = 'vertical')

    # now determine nice limits by hand:
    binwidth = 0.5
    xmax = max(x)
    xmin = min(x)
    #xymax = max(np.max(np.abs(x)), np.max(np.abs(y)))
    #lim = (int(max/binwidth) + 1) * binwidth

    bins = np.arange(xmin-binwidth, xmax + binwidth, binwidth)
    ax_histx.hist(x, bins=bins)
    plt.ylabel("frequency")
    #ax_histy.hist(y, bins=bins, orientation='horizontal')
#%% (4) start sequence

for elem in elems:
    
    target_var = targetVariableType + elem 
    
    
    """
    (4a) Load Unscaled Data
    """
    
   # training data 

    ytrain = df_train[target_var] #target col name GDOES_saltConc_Fe_mols
    Xtrain_scaled = df_train_scaled.drop(['RandomNo','PlateID', 
                                     'Sample#',
                                     targetVariableType+'Cr',
                                     targetVariableType+'Fe',
                                     targetVariableType+'Mn',
                                     targetVariableType+'Ni',
                                     targetVariableType+'total'],axis='columns')#
    numSamples_train = df_train.shape[0]
    #Xprune = df_full[['x(Cr)','x(Fe)','x(Mn)','x(Ni)','Fe/Mn','Cr/Mn','(Cr+Fe)/Mn']]
    #Xprune = X.drop(['x(Cr)','x(Fe)','x(Mn)','x(Ni)','Fe/Mn','Cr/Mn','(Fe+Cr)/Mn'])
    ytrain_exp = ytrain
    #y = (yUnscaled - yUnscaled.mean()) / (yUnscaled.max() - yUnscaled.min())
    Xtrain_unpruned_exp = Xtrain_scaled#randomness feature is always the 1st feature in the X-matrix
    #normScale.fit(XUnscaled)
    #X = normScale.fit_transform(XUnscaled)

    #%% (4b) Fit Unpruned
    """
    (4b) Get List of Features for everyone. We assume the cols are exactly the same for X_train & X_test. 
    that's why you format the cols. first using the mastercopy then split only on rows. 
    """
    X_len = len(Xtrain_unpruned_exp)#number of samples
    X_names_unpruned = Xtrain_unpruned_exp.columns #Xnames
    X_names_unpruned_arr = np.transpose(np.array([X_names_unpruned]))
    numFeats_unpruned = np.size(X_names_unpruned_arr)
    
    model_randFor = RandomForestRegressor(n_estimators=1000, random_state = 1,verbose=0) #initialize hyperparams for model_randfor
    unpruned_train_Rsq,unpruned_train_y_fit,unpruned_train_RMSE_fit,unpruned_train_uncert_down_y_predict_np, unpruned_train_uncert_up_y_predict_np, unpruned_train_uncertMargin_y_pred, unpruned_train_yfit_xvalPredict,unpruned_train_RMSE_xval,unpruned_train_yfit_percPredErrs,unpruned_train_yfit_xvalMAPE,unpruned_train_yfit_xvalMAPE_stdev,unpruned_train_featImpts=MLtrainsetMetrics(model_randFor,Xtrain_unpruned_exp,ytrain_exp)
    

    
                                                                                                                                                                                                                                                                                                                                                                                                                                          
#     #%% (4b) Feature pruning procedure based on correlations
#     """ 
#     (4b) Rand For for All Unscaled & Calculate KPI: xval m%e
                                         
#     """    
#     #initialize matrices first
#     X5_reduced = Xtrain_unpruned_exp    
#     X5_spear_corrMatrix = Xtrain_unpruned_exp.corr(method='spearman').abs()
#     recordCorrMatrix = X5_spear_corrMatrix.copy()#you have to do it this way or else when you do the "replace value with name" you will turn these to strings too
#     recordCorrMatrix_reduced = X5_spear_corrMatrix.copy()
#     X5_spear_reduced = X5_spear_corrMatrix.ge(threshold_spear).sum()
#     X5_spear_reduced_arr = np.array(X5_spear_reduced)
#     numFeats_reduced = numFeats_unpruned
#     # #%%
#     # for name, values in recordCorrMatrix.iteritems():
#     #     recordCorrMatrix.loc[recordCorrMatrix[name] > Spear_threshold, name] = str(name)
#     # recordCorrMatrix = recordCorrMatrix.T
#     # #^this line work
#     while X5_spear_reduced.max() >1: #while there are still features that have spear>0.85 with other features
#         hiCorrelFeatures = np.where(X5_spear_reduced_arr==X5_spear_reduced.max()) #find the indices of the ones with the highest number of spear>0.85
#         print(hiCorrelFeatures)#the features picked out here are the features that are 
#         print(recordCorrMatrix.columns[hiCorrelFeatures])
#         # for k in np.nditer(hiCorrelFeatures): #range(numFeats):  # for each of the high-corr features  
#         #     k = int(k) #ok this loop only bags features that are eliminated in each round. it does not bag all features 
#         #     #print(k)
#         #     colName = recordCorrMatrix.columns[k] #get its name
#         #     #print(colName)
#         #     #print(colName)
#         #     for j in range(numFeats_unpruned):
#         #         #print(j)
#         #         try: 
#         #             if recordCorrMatrix.iloc[k,j] > threshold_spear: #cycle through the rows, containing ALL ORIGINAL FEATURES
#         #                 recordCorrMatrix.iloc[k,j] = str(colName)
#         #         except:
#         #             pass
                
#         # for k in np.nditer(hiCorrelFeatures): #range(numFeats):  # for each of the high-corr features  #np.nditer(hiCorrelFeatures):
#         #     k = int(k)
#         #     #print(k)
#         #     colName = recordCorrMatrix.columns[k] #get its name
#         #     #print(colName)
#         #     #print(colName)
#         #     for j in range(numFeats_reduced):
#         #         #print(j)
#         #         try: 
#         #             if recordCorrMatrix_reduced.iloc[k,j] > Spear_threshold: #cycle through the rows, containing ALL ORIGINAL FEATURES
#         #                 recordCorrMatrix_reduced.iloc[k,j] = str(colName)
#         #         except:
#         #             pass
        
#         X5_reduced = X5_reduced.drop(X5_reduced.columns[hiCorrelFeatures],axis=1) #drop these hi-corr features from the input matrix
#         numSampls, numFeats_reduced = X5_reduced.shape #find the new number of features 
#         #recordCorrMatrix_reduced = 
#         recordCorrMatrix_reduced = recordCorrMatrix_reduced.drop(recordCorrMatrix_reduced.columns[hiCorrelFeatures],axis=1) #drop these features from the spearman correlaton coefficient recorder matrix
#         X5_correl_reduced = X5_reduced.corr(method='spearman').abs()#this is the spearman reduced matrix, still a square
# 		#print spearman reduced matrix 
#         X5_spear_reduced = X5_correl_reduced.ge(threshold_spear).sum()
#         X5_spear_reduced_arr = np.array(X5_spear_reduced)
#         print(X5_spear_reduced.max())
                                               

        
#     for j in range(numFeats_unpruned):#rows
#         colName = recordCorrMatrix.columns[j]
#         for i in range(numFeats_unpruned):#cols
#             try:
#                 if recordCorrMatrix.iloc[i,j] > threshold_spear:
#                     recordCorrMatrix.iloc[i,j] = colName
#             except:
#                 pass
        
#     #%%
#     """
#     ok have a block here that takes recordCorrMatrix,
#     and in every row, counts which of the retained features appear.
#     bin this with each retained feature. 
#     do this for all 63 rows
#     """
                                                                                                                                                                                                                                                                                                                                                                                                                                                                      
    
#     featureGroupsRecord = Workbook()
#     ws1 = featureGroupsRecord.active
#     ws1.title = "AllFeatureGroups"
#     ws1.append(['GroupID','GroupLeader','Followers']) #header
#     #activSheet = 
    
    
#     numRows_recordCorrMatrix, numCols_recordCorrMatrix = recordCorrMatrix.shape
#     rowIndices_recordCorrMatrix = []
#     for rowIndex in range(numRows_recordCorrMatrix):
#         rowName = recordCorrMatrix.index[rowIndex]
#         rowIndices_recordCorrMatrix.append(rowIndex)
#         shellGroup = ['Group ' + str(rowIndex)] #dummy variable in which to save the matrix while it's iterating. it is cleared every round
#         shellGroup.append(rowName)
#         for colIndex in range(numCols_recordCorrMatrix):
#             if type(recordCorrMatrix.iloc[rowIndex,colIndex])== str:
#                 shellGroup.append(recordCorrMatrix.iloc[rowIndex,colIndex])
        
#         exec(f'group_{colIndex} = {shellGroup}') #convert these to a dataframe and export to excel
#         ws1.append(shellGroup)
    
#     featureGroupsRecord.save(resultsExportDir + featureGroupRecord)
#     Xtrain_spearPruned_exp = X5_reduced
    
#     X_names_reduced = X5_reduced.columns
#     X_names_reduced_arr = np.transpose(np.array([X_names_reduced]))
#     X_len_reduced = len(X_names_reduced)
#     featureBins = pd.read_excel(resultsExportDir + featureGroupRecord,sheet_name = 'AllFeatureGroups')
    
#     ws2 = featureGroupsRecord.create_sheet("Sheet2")
#     ws2.title = "SpearPruneFeatures"
#     ws2.append(['GroupID','GroupLeader','Followers']) #header
#     #activSheet = featureGroupsRecord.active
    
#     for i in range(X_len_reduced):
#         featureOfInterest = X_names_reduced[i]
#         print(featureOfInterest)
#         featRow = featureBins[featureBins['GroupLeader']==featureOfInterest]
#         print(featRow)
#         featRow_list  = featRow.values.tolist()[0]#when you do pd.values.tolist() it becomes a list of lists, so you have to extract the "first" (and only) list from the list of list to make it a single list.
#         print(featRow_list)
#         ws2.append(featRow_list)
        
#     featureGroupsRecord.save(resultsExportDir + featureGroupRecord)
#     #%%
    """
    ok have a block here that takes recordCorrMatrix,
    and in every row, counts which of the retained features appear.
    bin this with each retained feature. 
    do this for all 63 rows
    """
    #%% (4c) Fit spearPruned Random Forest
    model_randFor = RandomForestRegressor(n_estimators=1000, random_state = 1) #initialize hyperparams for model_randfor
    Xtrain_spearPruned_exp = Xtrain_unpruned_exp
    spearPruned_train_Rsq,spearPruned_train_y_fit,spearPruned_train_RMSE_fit,spearPruned_train_uncert_down_y_predict_np, spearPruned_train_uncert_up_y_predict_np, spearPruned_train_uncertMargin_y_pred, spearPruned_train_yfit_xvalPredict,spearPruned_train_RMSE_xval,spearPruned_train_yfit_percPredErrs,spearPruned_train_yfit_xvalMAPE,spearPruned_train_yfit_xvalMAPE_stdev,spearPruned_train_featImpts=MLtrainsetMetrics(model_randFor,Xtrain_unpruned_exp,ytrain_exp)
    
    # #%% (4d) Rand Seed Perturb Pruner
    # for i in range(randSeedPerturbReps):
    #     print(i)
    #     randForRandSeed = np.random.randint(low=1,high=1000)
    #     X_len_spearPruned = len(Xtrain_spearPruned_exp)
    #     X_names_spearPruned = Xtrain_spearPruned_exp.columns
    #     X_names_spearPruned_arr = np.transpose(np.array([X_names_spearPruned]))
    #     X_names_spearPruned_df = pd.DataFrame([X_names_spearPruned])
        
    #     model_randForRandPert = RandomForestRegressor(n_estimators=1000, random_state = randForRandSeed)
    #     mod_ft_impts_scrs=MLtrainsetMetrics(model_randForRandPert,Xtrain_spearPruned_exp,ytrain_exp)[11]
        
    #     ftRankNos = np.argsort(mod_ft_impts_scrs)[::-1] # this pulls out a vector of the imptce ranking # of Xnames vector.
        
    #     if i ==0:
    #         ftRankNos_collected = ftRankNos
    #     else:
    #         ftRankNos_collected = np.column_stack((ftRankNos_collected,ftRankNos))

    # """
    # At the end of the Rand Seed perturbation, collect the mean and std of the feature rank nos.
    # """
    # imptsRankPerFeature_average =np.mean(ftRankNos_collected,axis=1)
    # imptsRankPerFeature_average_df =pd.DataFrame([imptsRankPerFeature_average])
    # imptsRankPerFeature_average_df = imptsRankPerFeature_average_df.T
    # X_names_spearPruned_df = X_names_spearPruned_df.T
    # imptsRankPerFeature_stdev = np.std(ftRankNos_collected,axis=1)#np, axis=0 --> operation along the columns
    
    # aveFtRanks_df = pd.DataFrame(np.column_stack((X_names_spearPruned_arr[:,0],imptsRankPerFeature_average,imptsRankPerFeature_stdev)))
    # aveFtRanks_df_sorted = aveFtRanks_df.sort_values(1,ascending=True)
    # randomFeatRow = aveFtRanks_df_sorted[aveFtRanks_df_sorted[0].str.contains("RandomNo")]#pull out the average rank for the random feature
    # randomFeatRank = randomFeatRow.iat[0,1]
    # aveFtRanks_np_sorted = np.array(aveFtRanks_df_sorted)
    # rankMargin = 0
    # counter = 0
    # FEATURESofInterest2Plot = np.array(randomFeatRow)
    # while rankMargin < randomFeatRank:
    #     rankMargin = aveFtRanks_np_sorted[counter,1] #+ aveFtRanks_np_sorted[counter,2]
    #     FEATURESofInterest2Plot = np.row_stack((FEATURESofInterest2Plot,aveFtRanks_np_sorted[counter,:])) #vstack?
    #     counter = counter+1
    # #%%
    # """
    # ultra-reduced feature space to fit a randfor
    # Redefine feature vector, do new model fit, then test
    # """
    # X_doublePruned = pd.DataFrame()
    # colIndices_featureDublPruned = []
    # featureBins = pd.read_excel(resultsExportDir + featureGroupRecord)
    
    # ws3 = featureGroupsRecord.create_sheet("Sheet3")
    # ws3.title = "RandSeedPertSelected"
    # ws3.append(['GroupID','GroupLeader','Followers']) #header
    
    # for i in range(counter+1):
    #     #this segment prepares the X_dublPruned matrix for fitting RFR
    #     featureOfInterest = FEATURESofInterest2Plot[i,0]
    #     colIndex_feature = Xtrain_unpruned_exp.columns.get_loc(featureOfInterest)
    #     data_featureOfOInterest = Xtrain_spearPruned_exp[[featureOfInterest]]
    #     X_doublePruned = pd.concat([X_doublePruned, data_featureOfOInterest],axis = 1)
    #     #this segment prepares the labels for the importance ranking barplot
    #     if i == 0:
    #         colIndices_featureDublPruned = ['Group ' + str(colIndex_feature)]
    #         randSeedPerturbSavedFeats = pd.DataFrame()
    #     else:
    #         colIndices_featureDublPruned.append('Group ' + str(colIndex_feature))
    #     #this segment will select the feature groups from the excel sheet
    #     featRow = featureBins[featureBins['GroupLeader']==featureOfInterest]
    #     # #randSeedPerturbSavedFeats = pd.concat([randSeedPerturbSavedFeats,featRow])#this works
    #     featRow_list  = featRow.values.tolist()[0]#when you do pd.values.tolist() it becomes a list of lists, so you have to extract the "first" (and only) list from the list of list to make it a single list.
    #     print(featRow_list)
    #     ws3.append(featRow_list)
        
    # featureGroupsRecord.save(resultsExportDir + featureGroupRecord)
    # X_doublePruned = X_doublePruned.drop(['RandomNo'],axis='columns')#
    # randSeedPerturbSavedFeats.to_excel(resultsExportDir + featureGroupRecord,sheet_name='RandSeedPertSelected')  
        

    # selectedUncorrFeatures = X_doublePruned.columns
    # X_len_dublPruned = len(X_doublePruned)
    # X_names_dublPruned = X_doublePruned.columns
    # X_names_dublPruned_arr = np.transpose(np.array([X_names_dublPruned]))
    # X_names_dublPruned_df = pd.DataFrame([X_names_dublPruned])
    # colIndices_featureDublPruned_arr = np.array(colIndices_featureDublPruned)
    
    # dublPruned_train_Rsq, dublPruned_train_y_fit,dublPruned_train_RMSE_fit,dublPruned_train_uncert_down_y_predict_np, dublPruned_train_uncert_up_y_predict_np, dublPruned_train_uncertMargin_y_pred, dublPruned_train_yfit_xvalPredict,dublPruned_train_RMSE_xval,dublPruned_train_yfit_percPredErrs,dublPruned_train_yfit_xvalMAPE,dublPruned_train_yfit_xvalMAPE_stdev,dublPruned_train_featImpts=MLtrainsetMetrics(model_randFor,X_doublePruned,ytrain_exp)

    #%% (6a) Scatterplot: fit
    plt.rcParams['figure.figsize'] = [6, 6]
    fig1 = plt.figure()
    plt.plot(ytrain_exp,ytrain_exp,color='blue',label='parity plot')#doing plt.plot or plt.scatter in succession overlays the plot by default.
    #feature importances for the model of each elem
    randForUncertBand_df = pd.DataFrame(np.column_stack((ytrain_exp,unpruned_train_uncert_up_y_predict_np,unpruned_train_uncert_down_y_predict_np)))
    randForUncertBand_df_sorted = randForUncertBand_df.sort_values(0,ascending=True)#1st arg is the col which you're sorting by
    randForUncertBand_arr_sorted = np.array(randForUncertBand_df_sorted)
    plt.fill_between(randForUncertBand_arr_sorted[:,0], randForUncertBand_arr_sorted[:,1],randForUncertBand_arr_sorted[:,2], alpha=0.2, linewidth=0, label = 'random forest uncertainty band')# ' twice-feature pruned')
    plt.scatter(ytrain_exp,unpruned_train_y_fit,label='unpruned training set ('+str(numSamples_train)+')',edgecolors='green',c = 'none')#c = 'green') #
    plt.scatter(ytrain_exp,spearPruned_train_y_fit,label='Spear-prune val:'+str(threshold_spear)+' train set ('+str(numSamples_train)+')',color='green')#, alpha = 0.7)
    #plt.scatter(ytrain_exp,dublPruned_train_y_fit,label='randSeed+Spearman-pruned training set ('+str(numSamples_train)+')',color='green')
    # plt.scatter(ytest_goodCorr_exp,ytest_goodCorr_pred, label='test set selected' , color = 'red')
    # plt.scatter(ytest_exp,y_test_predict, label='test set ('+str(numSamples_test)+')',color='red')
    # plt.scatter(ytest_badCorr_exp,ytest_badCorr_pred, label='test set rejected', color = 'black')
    plt.title(" Target: "+ elem)
    plt.xlabel("Experimental value" + unit)
    plt.ylabel("Random Forest predicted value" + unit)
    
    # plot n save the scatter plot
    plt.legend()
    plt.show()#displays the plot in the console
    parityPlotName = modelType+"_"+elem+"_RFparityPlot.png"
    fig1.savefig(resultsExportDir + parityPlotName)#save plot
    
    # #%% (6a) Scatterplot: xval
    # plt.rcParams['figure.figsize'] = [6, 6]
    # fig6 = plt.figure()
    # plt.plot(ytrain_exp,ytrain_exp,color='blue',label='parity plot')#doing plt.plot or plt.scatter in succession overlays the plot by default.
    # #feature importances for the model of each elem
    # randForUncertBand_df = pd.DataFrame(np.column_stack((ytrain_exp,unpruned_train_uncert_up_y_predict_np,unpruned_train_uncert_down_y_predict_np)))
    # randForUncertBand_df_sorted = randForUncertBand_df.sort_values(0,ascending=True)#1st arg is the col which you're sorting by
    # randForUncertBand_arr_sorted = np.array(randForUncertBand_df_sorted)
    # plt.fill_between(randForUncertBand_arr_sorted[:,0], randForUncertBand_arr_sorted[:,1],randForUncertBand_arr_sorted[:,2], alpha=0.2, linewidth=0, label = 'random forest uncertainty band twice-feature pruned')
    # plt.scatter(ytrain_exp,unpruned_train_yfit_xvalPredict,label='unpruned training set ('+str(numSamples_train)+')',edgecolors='green',c = 'none')
    # plt.scatter(ytrain_exp,spearPruned_train_yfit_xvalPredict,label='Spearman-pruned training set ('+str(numSamples_train)+')',color='green', alpha = 0.7)
    # plt.scatter(ytrain_exp,dublPruned_train_yfit_xvalPredict,label='randSeed+Spearman-pruned training set ('+str(numSamples_train)+')',color='green')
    # # plt.scatter(ytest_goodCorr_exp,ytest_goodCorr_pred, label='test set selected' , color = 'red')
    # # plt.scatter(ytest_exp,y_test_predict, label='test set ('+str(numSamples_test)+')',color='red')
    # # plt.scatter(ytest_badCorr_exp,ytest_badCorr_pred, label='test set rejected', color = 'black')
    # plt.title("Model type: " + modelType + " Target: "+ elem)
    # plt.xlabel("Experimental value" + unit)
    # plt.ylabel("Random Forest XVAL predicted value" + unit)
    
    # # # this segment is for the confusion matrix
    # # plt.hlines(SS316CorrCriteria,0,3500,linestyles='dashed')
    # # plt.vlines(SS316CorrCriteria,0,3500,linestyles='dashed')
    # # plt.text(0,0,"true positives")
    # # plt.text(SS316CorrCriteria,0,"false positives")
    # # plt.text(SS316CorrCriteria,2400,"true negatives")
    # # plt.text(0,2400,"false negatives")
    # # plt.text(SS316CorrCriteria,SS316CorrCriteria,"y test MAPE = " + str(round(ytestfit_MAPE,0)) + "+/-" + str(round(ytestfit_MAPE_stdev,0)) + " %")
    
    # # plot n save the scatter plot
    # plt.legend()
    # plt.show()#displays the plot in the console
    # parityPlotName = modelType+"_"+elem+"_RFparityPlot.png"
    # fig6.savefig(resultsExportDir +" XVAL "+  parityPlotName)#save plot
    
    # #%% (6b) 2-prune % pred err histogram
    # plt.rcParams['figure.figsize'] = [6.4, 4]
    # fig2 = plt.figure()
    # plt.hist(dublPruned_train_yfit_percPredErrs,edgecolor='black', linewidth=1.2)#%errs
    # plt.title("Double-pruned train % pred errs, Target: "+ elem)
    # plt.figtext(0.6,0.8,"Mean abs%err= "+str(round(dublPruned_train_yfit_xvalMAPE,1))+"%") #the x,y coords represent "relative proportion of xy axes".
    # plt.xlabel("abs % err between Model Prediction - Experimental")
    # plt.ylabel("No. of Samples")
    # hisPlotName = modelType+"_"+elem+"_%errHist.png"
    # fig2.savefig(resultsExportDir +hisPlotName)
    # #%% (6c) Importance score printing module for 2-pruned feature set
    # """
    # (7) Data Printing module
    # """

    # mod_ft_impts_scrs = pd.DataFrame(dublPruned_train_featImpts,columns = ['imptsScore']) # model_randFor.feature_importances_
    # mod_ft_impt_feats = pd.DataFrame(X_names_dublPruned,columns=['feature'])
    # mod_ft_impts_scrs_df = pd.concat([mod_ft_impt_feats,mod_ft_impts_scrs],axis = 1)
    # mod_ft_impts_scrs_df_sorted = mod_ft_impts_scrs_df.sort_values(by='imptsScore',ascending=False)#1st arg is the col which you're sorting by
    # #append using np but convert to df for printing as csv.
    # imp_scrs_filename = "ImptScrs"+elem+".csv"
    # mod_ft_impts_scrs_df_sorted.to_csv(resultsExportDir + modelType + imp_scrs_filename)
    
    # #%% (6d)Plotting the twice-reduced feature set: once from Spearman Prune, next from random seed perturbation prune
    # featureGroupsRecord.save(resultsExportDir + featureGroupRecord)
    # fig4,ax4 = plt.subplots()
    
    # FEATURESofInterest2Plot_droplastrow = FEATURESofInterest2Plot[:-1]
    # FEATURESofInterest2Plot_droplastrow_df = pd.DataFrame(FEATURESofInterest2Plot_droplastrow, columns = ['Feature','ave rank','rank stdev'])
    # FEATURESofInterest2Plot_droplastrow_df.to_excel(resultsExportDir + 'DoublePruneAveRankings.xlsx')
    # xpos = np.arange(FEATURESofInterest2Plot_droplastrow.shape[0])
    # ax4.bar(xpos,FEATURESofInterest2Plot_droplastrow[:,1],yerr=FEATURESofInterest2Plot_droplastrow[:,2],align='center', alpha=0.5, ecolor='black', capsize=10)
    # ax4.set_ylabel('Ordinal Importance Ranking')
    # ax4.set_xticks(xpos)
    # #ax3.set_xticklabels(FEATURESofInterest2Plot[:,0], rotation = 45, ha="right")
    # labels = ['Randomness feature'] #randomness feature is always the 1st feature in the X-matrix
    # # for i in range(1,counter+1):
    # #     labels.append('Group ' + str(i))
    # colIndices_featureDublPruned_droplastrow = colIndices_featureDublPruned[:-1]
    # ax4.set_xticklabels(FEATURESofInterest2Plot_droplastrow[:,0], rotation = 45, ha="right")
    # #ax4.set_xticklabels(colIndices_featureDublPruned_droplastrow, rotation = 45, ha="right")
    # plt.xticks(fontsize=8)
    # plt.title("Importance Ranking Perturbation wRandSeed" + " Target:"+ elem)
    # ax4.yaxis.grid(True)
    # plt.tight_layout()
    # plt.ylim([0,randomFeatRank+5])
    # #plt.xlim([-1, numFeatsOfInterest]) #plot top 10 features only because it's a bit much
    # fig4.savefig(resultsExportDir + "randSeedPerturb_ImptcRank.png")
    
    # #%% (6e) Comparison of mape
    # fig5,ax5 = plt.subplots()
    # xpos_fig5 = np.arange(3)
    
    # d = {'labels': ['unpruned', 'Spearman-pruned feature set', 'Spear/Randseed-pruned feat set'],'MAPE': [unpruned_train_yfit_xvalMAPE,spearPruned_train_yfit_xvalMAPE,dublPruned_train_yfit_xvalMAPE], 'stdev': [unpruned_train_yfit_xvalMAPE_stdev, spearPruned_train_yfit_xvalMAPE_stdev,dublPruned_train_yfit_xvalMAPE_stdev]}
    # MAPEs_df = pd.DataFrame(data=d)
    # ax5.bar(xpos_fig5,MAPEs_df['MAPE'],yerr = MAPEs_df['stdev'],align='center', alpha=0.5, ecolor='black', capsize=10)
    # ax5.set_ylabel('Comparing MAPE for feature set pruning')
    # ax5.set_xticks(xpos_fig5)
    # ax5.set_xticklabels(MAPEs_df['labels'], rotation = 45, ha="right")
    # plt.xticks(fontsize=8)
    # plt.title("Comparing MAPE for feature set pruning" + " Target:"+ elem)
    # ax5.yaxis.grid(True)
    # plt.tight_layout()
    # fig5.savefig(resultsExportDir+"comparingFeaturePruning.png")
    
    
        
    #%%
    """
    (7a) Shapley calculation
    """
    Z_model_randFor_fitted = model_randFor.fit(Xtrain_spearPruned_exp,ytrain_exp)
    
    
    #Z_shapSplainer = shap.Explainer(Z_model_randFor_fitted)
    #Z_shapSplanation = Z_shapSplainer(Xtrain_spearPruned_exp)
    #Z_shapValues_literal = Z_shapSplainer.shap_values(Xtrain_spearPruned_exp)

    Z_shapSplainer_rfr = shap.explainers.Tree(Z_model_randFor_fitted)#work
    Z_shapSplanation_rfr = Z_shapSplainer_rfr(Xtrain_spearPruned_exp)#work
    Z_shapValues_rfr_literal = Z_shapSplainer_rfr.shap_values(Xtrain_spearPruned_exp)
    Z_shapValues_rfr_literal = pd.DataFrame(Z_shapValues_rfr_literal,columns = X_names_unpruned)
    #%%
    inds = np.argsort(np.abs(Z_shapSplanation_rfr.values).mean(0))#this is the order of most important to least important
    # for i in inds:
    #     scatter_hist(Xtrain_spearPruned_exp.iloc[:,i], Z_shapValues_rfr_literal[:,i],X_names_reduced[i])
    
    #shap.plots.scatter(Z_shapSplanation[:,0])#works
    #shap.plots.scatter(Z_shapSplanation_rfr[:,0])#works
    
    # clust = shap.utils.hclust(Xtrain_spearPruned_exp,ytrain_exp, linkage="single")
    # from scipy.cluster import hierarchy
    # plt.rcParams['figure.figsize'] = [6.4, 4]
    # #fig2 = plt.figure()
    # hierarchy.dendrogram(clust,labels=Xtrain_spearPruned_exp.columns, leaf_rotation=90)# for how to interpret this see: https://docs.scipy.org/doc/scipy/reference/generated/scipy.cluster.hierarchy.dendrogram.html?highlight=dendrogram
    shap.summary_plot(Z_shapSplanation_rfr, Xtrain_spearPruned_exp, show=False,use_log_scale=False)
    shap.plots.bar(Z_shapSplanation_rfr, max_display=18)

    # Z_shapValues_rfr_literal_minusDssNi = np.delete(Z_shapValues_rfr_literal,29,axis=1)
    # Xtrain_spearPruned_exp_minusDssNi = Xtrain_spearPruned_exp.drop(['$D_{SS,Ni}$'],axis='columns')
    # shap.summary_plot(Z_shapValues_rfr_literal_minusDssNi, Xtrain_spearPruned_exp_minusDssNi) #this also works too and it's more understandable
    
    # selected
    selectedFeatureIndex = [0,1,2,3,4,5,7,11,12,14,22,23,25,26,27,28,29,30]
    Z_shapValues_rfr_literal_abs = abs(Z_shapValues_rfr_literal)
    Z_shapValues_rfr_literal_absMean = Z_shapValues_rfr_literal_abs.mean()
    Z_shapValues_rfr_literal_absMeanSorted = Z_shapValues_rfr_literal_absMean.sort_values(0,ascending=False)
    Z_shapValues_rfr_literal_absMeanSorted_spearSelect = Z_shapValues_rfr_literal_absMeanSorted[selectedFeatureIndex]
    #ok make the barplot here
    import seaborn as sns
    import matplotlib.colors
    cmap = plt.cm.rainbow
    norm = matplotlib.colors.Normalize(vmin=min(Z_shapValues_rfr_literal_absMean), vmax=266)
    Z_shapValues_rfr_literal_absMean_np = np.array(Z_shapValues_rfr_literal_absMean)
    Z_shapValues_rfr_literal_absMean_np = np.transpose(Z_shapValues_rfr_literal_absMean_np)
    
    Z_shapValues_rfr_literal_absMean_matr = pd.DataFrame()
    #%%
    for i in np.arange(1,110):
        Z_shapValues_rfr_literal_absMean_matr = pd.concat([Z_shapValues_rfr_literal_absMean_matr,Z_shapValues_rfr_literal_absMean],axis='columns')
    #%%
    Z_shapValues_rfr_literal_absMean_matr = Z_shapValues_rfr_literal_absMean_matr.T
    chartStrip = sns.stripplot(data=Xtrain_unpruned_exp, palette = "cool", orient = "h", size = 5 ,alpha=0.5,)
    chartSwarm = sns.swarmplot(data=Xtrain_unpruned_exp, orient = "h", size = 1)
    
    
    #shap.plots.bar(Z_shapSplanation,clustering=clust,clustering_cutoff=1, show=False)#you have to put show=False or it won't save
    #https://shap.readthedocs.io/en/latest/example_notebooks/api_examples/plots/bar.html
    #xplainer = shap.TreeExplainer(model)
    expected_value = Z_shapSplainer_rfr.expected_value
    if isinstance(expected_value, list):
        expected_value = expected_value[1]
    print(f"Explainer expected value: {expected_value}")
    
    select = range(20)
    feature = Xtrain_spearPruned_exp.iloc[select]
    feature_display = Xtrain_spearPruned_exp.loc[feature.index]
    
    # with warnings.catch_warnings():
    #     warnings.simplefilter("ignore")
    shap_values = Z_shapSplainer_rfr.shap_values(feature)[1]
    shap_interaction_values = Z_shapSplainer_rfr.shap_interaction_values(feature)
    if isinstance(shap_interaction_values, list):
        shap_interaction_values = shap_interaction_values[1]
   
    shap.decision_plot(expected_value, Z_shapValues_rfr_literal[1,:], feature_display)#this works# https://shap.readthedocs.io/en/latest/example_notebooks/api_examples/plots/decision_plot.html
    
    #works up to this point
    plt.savefig(resultsExportDir+"shapSummary.png")
    
    clust = shap.utils.hclust(Xtrain_spearPruned_exp,ytrain_exp, linkage="single")
    plt.savefig(resultsExportDir+"shapvalues.png")
    
    # explainer.expected_value = explainer.expected_value[0]  # Additional line to force the dim of base_values
    # rf_shap_values = explainer(X_doublePruned)
    # shap.plots.scatter(rf_shap_values[:,"RM"], color=rf_shap_values)
    
    shap.plots.scatter(shap_values, ylabel="SHAP value\n(higher means greater corrosion)", show=False)
    plt.savefig(resultsExportDir+"shapValues.png")


    #%%
    DTDC_Mn = Xtrain_scaled[['DC(Mn,Mn)','DT(Mn)']]
    DTDC_Mn.corr(method='spearman')
    
    #%%
    
    for featInd in np.arange(numFeats_unpruned):
        feat_vs_shap = pd.DataFrame([Xtrain_scaled.iloc[:,featInd], Z_shapValues_rfr_literal.iloc[:,featInd]])
        feat_vs_shap = np.transpose(feat_vs_shap)
        featureShapSpearVal = feat_vs_shap.corr(method='spearman').iloc[0,1]
        feature_N_itsShapSpearVal = {X_names_unpruned[featInd]:[featureShapSpearVal]}
        if featInd ==0:
            features_N_theirShapSpearVals = pd.DataFrame(feature_N_itsShapSpearVal)
        else: 
            feature_N_itsShapSpearVal = pd.DataFrame(feature_N_itsShapSpearVal)
            features_N_theirShapSpearVals = pd.concat([features_N_theirShapSpearVals,feature_N_itsShapSpearVal],axis='columns')
    
    #features_N_theirShapSpearVals = features_N_theirShapSpearVals.transpose()
    #features_N_theirShapSpearVals shows the spearman coeff of the shapley. indicates the strength of trend of the feature. 
    Z_shapValues_rfr_literal_abs = Z_shapValues_rfr_literal.abs()
    Z_shapValues_rfr_literal_abs_mean = np.transpose(Z_shapValues_rfr_literal_abs.mean())#indicates the average impact of feature
    
    shaps = np.array(Z_shapValues_rfr_literal_abs_mean)
    spears_abs = features_N_theirShapSpearVals.abs()
    spears_abs = spears_abs.values.flatten()
    spears =  features_N_theirShapSpearVals.values.flatten()
    Shaps_vs_Spears = df = pd.DataFrame({'shaps': shaps,'spears_abs': spears_abs}, index=X_names_unpruned)
    Shaps_vs_Spears_sorted = Shaps_vs_Spears.sort_values(by='spears_abs',ascending=False)
    Shaps_vs_Spears_sorted_feats = Shaps_vs_Spears_sorted.columns
    hiSpearShaps= Shaps_vs_Spears_sorted.iloc[0:19,0]
    hiSpearShaps_shapSorted = hiSpearShaps.sort_values(ascending=True)
    # hiSpearShaps = Shaps_vs_Spears_sorted.sort_values(by='shaps', ascending=False)
    plt.rcParams['figure.figsize'] = [6,12]
    ax = hiSpearShaps_shapSorted.plot.barh(xlabel='Feature',title='Mean |SHAP| per feature',width=0.8, fontsize=16, sort_columns=True)#,log=True
    #%%
    fig_shapsVspears = plt.figure() # Create matplotlib figure
    ax_shapsVspears = fig_shapsVspears.add_subplot(111) # Create matplotlib axes
    #ax_shapsVspears_2 = ax_shapsVspears.twinx() # Create another axes that shares the same x-axis as ax.
    width = 0.4
    Shaps_vs_Spears.shaps.plot(kind='bar', color='blue', ax=ax_shapsVspears, width=width, position=1)
    #Shaps_vs_Spears.spears_abs.plot(kind='bar', color='orange', ax=ax_shapsVspears_2, width=width, position=0)
    ax_shapsVspears.set_ylabel('Mean Absolute Shapley value')
    #ax_shapsVspears_2.set_ylabel('Spearman coeff for Shapley values vs. Feature value')
    plt.show()
    
    
    #%%

executionTimeInMins = (time.time() - startTime)/60
print('Execution time in mins = ' + "{:.2f}".format(executionTimeInMins))
